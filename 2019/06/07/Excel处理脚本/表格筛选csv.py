# -*- coding:UTF-8 -*-
import os
import xlrd,xlwt,csv                                #需要pip安装xlrd，xlwt
from openpyxl import load_workbook                  #需要pip安装openpyxl
dir="C:/Users/Administrator/PycharmProjects/xmq"    #文件目录，根据自己的文件所在目录更改，将脚本放在表格的目录。


exclude_file=['hcx.csv','sx.csv','qc.csv']
file_type=['xlsx','csv']


def render_file(filename,data):
    '''
        写入文件
    :param filename:文件名
    :param data: 数据二维数组 [[]]
    :return:
    '''
    file = xlwt.Workbook()
    table = file.add_sheet('sheet1')
    for index,row in enumerate(data):
        for col_index,col in enumerate(row):
            table.write(index, col_index, col)
    file.save(filename)
    return filename
    
def collect_file():
    '''收集文件'''
    files=[]
    for file in os.listdir(dir):
        if file.split(".")[-1].lower() in file_type and file.split(".")[-1].lower() not in exclude_file:
            files.append(os.path.join(dir,file))
    return files

def collect_data():
    files=collect_file()
    #数据收集 每一行坐为一条记录 每条记录 用列表存储
    datas=[]
    for file in files:
        try:
            csv_file=csv.reader(open(file))
        except:
            raise ValueError("文件格式编码错误")
        for item in csv_file:
            datas.append(item)
        # # book = xlrd.open_workbook(file)
        # workbook = load_workbook(file)
        # sheets = workbook.get_sheet_names()
        # #如果有多个sheet  也需要存储
        # for sheet in sheets:
        #     booksheet=workbook.get_sheet_by_name(sheet)
        #     rows_count=booksheet.rows
        #     for row in range(rows_count):
        #         print(row)
        
    #生成合成后的文件
    render_file(os.path.join(dir,'hcx.csv'),datas)
    return datas
def filter_data(data):
    '''数据过滤 严格区分大小写'''
    if not data:
        print("没有数据源")
        return
    print("exp:%s"%"--".join([str(i) or "None" for i in data[0]]))
    col_index=input("请选择过滤列:1-%s"%len(data[0]))
    filter_index=int(col_index)-1
    if filter_index>=len(data[0]):
        #错误的列
        filter_data(data)
    _keys=input("请输入关键字,多个关键字用/隔开!")
    keys=_keys.split("/")
    d=[]
    for row in data:
        for key in keys:
            if key in str(row[filter_index]):
                d.append(row)
                break
    render_file("sx.csv",d)
    return d
    
def distict_data(data):
    '''数据去重  目前单列去重 如果需要多列去重 可以重复去重'''
    if not data:
        print("没有数据源")
        return
    print("exp:%s"%"--".join([str(i) or "None" for i in data[0]]))
    cnt1 = len(data)
    col_index = input("请选择去重列:1-%s" % len(data[0]))
    filter_index = int(col_index) - 1
    if filter_index >= len(data[0]):
        # 错误的列
        distict_data(data)
    keys=[]
    d=[]
    for item in data:
        key=item[filter_index]
        if key not in keys:
            keys.append(key)
            d.append(item)
    
    cnt2=len(d)
    print("去除重复行数:%s"%(cnt1-cnt2))
    render_file('qc.csv', d)
    return d



def show_action(data):
    ''' 功能项
    :param data:数据源
    :return:
    '''
    print("\033[32m1:文件过滤(默认:hcx.csv)")
    print("\033[35m2:文件去重(默认:sx.csv)")
    print("\033[35m输入Q退出")
    action=input("\033[36m请选择功能:")
    print(action)
    if action=="1":
        data=filter_data(data)
    elif action=="2":
        data=distict_data(data)
    elif action=="q" or action=='Q':
        return
    show_action(data)
    
def main():
    '''主函数  每次运行都需要读取数据到内存  过滤和去重直接操作该数据'''
    data=collect_data()
    show_action(data)
main()